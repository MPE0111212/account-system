import os.path
from tkinter import *
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
import openpyxl
import secrets

# Хеш и соль пароля аккаунта администратора, который создаётся при первом запуске (пароль: "admin")
ADMIN_PASSWORD = '0x6110991a231ab'
ADMIN_SALT = 'b45a1e9aa7086e93471c48590dd2e67f'

user_login = ''
user_clicks = 0

# Простая функция хеширования с шестнадцатиричными числами и солью для защиты от радужных таблиц
def trashcrypt(password, salt):
    salted_password = password + salt
    h = 89344537265732456
    for b in salted_password.encode('utf-8'):
        h ^= b
        h = (h * 109) & 0xFFFFFFFFFFFFF
    return hex(h)


# Проверка на существование БД
def check_for_databases():
    try:
        openpyxl.load_workbook('passwords.xlsx')
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['admin', ADMIN_PASSWORD, '*adm', ADMIN_SALT])
        wb.save('passwords.xlsx')


def register():
    try:
        wb = openpyxl.load_workbook('passwords.xlsx')
        ws = wb.active
        inp_login = login_V.get()
        inp_password = password_V.get()
        salt = str(secrets.token_hex(16))
        if not (2 < len(inp_login) < 21 and 2 < len(inp_password) < 21):
            warning_V.set("Логин и пароль должны быть длиннее 2 символов и короче 21 символа")
            return
        login_taken = False
        for i in ws.iter_rows():
            if i[0].value == inp_login:
                login_taken = True
                break
        if not login_taken:
            ws.append([inp_login, trashcrypt(inp_password, salt), '', salt])
            wb.save('passwords.xlsx')
            global user_login
            user_login = inp_login
            main_window()
        else:
            warning_V.set(f"Неверный логин или пароль")
    except:
        check_for_databases()
        user_login = login_V.get()
        main_window()


def login():
    global user_login, user_clicks
    try:
        wb = openpyxl.load_workbook('passwords.xlsx')
        ws = wb.active
        inp_login = login_V.get()
        login_exists = False
        account_clicks = 0
        password = ''
        salt = ''
        for i in ws.iter_rows():
            if i[0].value == inp_login:
                login_exists = True
                password = i[1].value
                if i[2].value:
                    account_clicks = i[2].value
                salt = i[3].value
                break
        if login_exists:
            if trashcrypt(password_V.get(), salt=salt) == password:
                user_login = inp_login
                user_clicks = account_clicks
                main_window()
            else:
                warning_V.set(f"Неверный логин или пароль")
        else:
            warning_V.set(f"Неверный логин или пароль")
    except:
        warning_V.set(f"В системе нет логинов")



def main_window():
    # Функции для администраторов
    def admin_full_reset():
        if os.path.exists('passwords.xlsx'):
            try:
                os.remove('passwords.xlsx')
                print("Removed passwords.xlsx")
            except:
                pass
        root2.destroy()

    def admin_register():
        try:
            wb = openpyxl.load_workbook('passwords.xlsx')
            ws = wb.active
            inp_login = admin_login_V.get()
            inp_password = admin_password_V.get()
            salt = str(secrets.token_hex(16))
            if not (2 < len(inp_login) < 21 and 2 < len(inp_password) < 21):
                admin_warning_V.set("Неверная длина логина или пароля")
                return
            login_taken = False
            for i in ws.iter_rows():
                if i[0].value == inp_login:
                    login_taken = True
                    break
            if not login_taken:
                ws.append([inp_login, trashcrypt(inp_password, salt), '*adm', salt])
                wb.save('passwords.xlsx')
                passwords_xlsx_elements = []
                wb = openpyxl.load_workbook('passwords.xlsx')
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    passwords_xlsx_elements.append(row)
                passwords_xlsx_text = ""
                for i in range(len(passwords_xlsx_elements)):
                    for j in range(len(passwords_xlsx_elements[i])):
                        passwords_xlsx_text += str(passwords_xlsx_elements[i][j]) + ' '
                    passwords_xlsx_text += '\n'
                passwords_xlsx_scrolled_text.delete('1.0', END)
                passwords_xlsx_scrolled_text.insert(INSERT, passwords_xlsx_text)
            else:
                admin_warning_V.set(f"Логин {inp_login} уже занят")
        except:
            check_for_databases()

    # Общие функции
    def click():
        global user_clicks
        user_clicks += 1
        clicks_V.set('Клики: ' + str(user_clicks))

    def save_and_quit():
        if is_admin:
            root2.destroy()
            return
        try:
            wb = openpyxl.load_workbook('passwords.xlsx')
            ws = wb.active
            login_row = -1
            r = 1
            for i in ws.iter_rows():
                if i[0].value == user_login:
                    login_row = r
                    break
                else:
                    r += 1
            if login_row != -1:
                ws[f'C{login_row}'] = user_clicks
                wb.save('passwords.xlsx')
                root2.destroy()
            else:
                ws.append([user_login, password_V.get(), user_clicks])
                wb.save('passwords.xlsx')
                root2.destroy()
        except:
            check_for_databases()
            root2.destroy()

    global root, user_clicks
    is_admin = user_clicks == '*adm'
    if is_admin:
        user_clicks = 0

    root.destroy()
    # Окно пользователя
    root2 = Tk()
    root2.attributes("-fullscreen", True)

    style2 = ttk.Style()
    style2.theme_use('alt')
    style2.configure('TEntry', fieldbackground='gray15')
    style2.configure('Click.TButton', background='turquoise1', foreground='turquoise4', font=("Arial", 40))
    style2.configure('Quit.TButton', background='red', foreground='black', font=("Arial", 15))
    style2.configure('Full_reset.TButton', background='red', foreground='black', font=("Arial", 10))
    style2.configure('Admin_register.TButton', background='RoyalBlue1', foreground='black', font=("Arial", 15))

    window2 = Label(background='gray10')
    window2.pack(fill=BOTH, expand=True, anchor='center')

    title2_label = ttk.Label(window2, text='СИСТЕМА АККАУНТОВ', font=("Arial", 30), anchor='center', background='gray10',
                             foreground='white')
    title2_label.place(relx=0, rely=0.05, relwidth=0.5, relheight=0.1)

    login2_label = ttk.Label(window2, text=user_login, font="Arial", anchor='center', background='gray15',
                             foreground='white')
    login2_label.place(relx=0.65, rely=0.05, relwidth=0.25, relheight=0.1)

    clicks_V = StringVar()
    clicks_V.set('Клики: ' + str(user_clicks))
    click_label = ttk.Label(window2, textvariable=clicks_V, font=("Arial", 30), anchor='center', background='gray5',
                            foreground='white')
    click_label.place(relx=0.25, rely=0.5, relwidth=0.5, relheight=0.1)

    click_button = ttk.Button(window2, text='Жмякать', style='Click.TButton', command=click)
    click_button.place(relx=0.35, rely=0.7, relwidth=0.3, relheight=0.15)

    quit2_button = ttk.Button(window2, text='Выйти', style='Quit.TButton', command=save_and_quit)
    quit2_button.place(relx=0.45, rely=0.075, relwidth=0.1, relheight=0.05)

    if is_admin:
        full_reset_button = ttk.Button(window2, text='УДАЛИТЬ БАЗУ ДАННЫХ\nИ ВЫЙТИ', style='Full_reset.TButton',
                                       command=admin_full_reset)
        full_reset_button.place(relx=0.02, rely=0.2, relwidth=0.1, relheight=0.1)

        passwords_xlsx_elements = []
        wb = openpyxl.load_workbook('passwords.xlsx')
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            passwords_xlsx_elements.append(row)
        passwords_xlsx_text = ""
        for i in range(len(passwords_xlsx_elements)):
            for j in range(len(passwords_xlsx_elements[i])):
                passwords_xlsx_text += str(passwords_xlsx_elements[i][j]) + ' '
            passwords_xlsx_text += '\n'
        passwords_xlsx_scrolled_text = ScrolledText(window2, background='gray8', foreground='white', font=("Arial", 10))
        passwords_xlsx_scrolled_text.place(relx=0.25, rely=0.2, relwidth=0.5, relheight=0.2)
        passwords_xlsx_scrolled_text.insert(INSERT, passwords_xlsx_text)

        admin_login_V = StringVar()
        admin_login_entry = ttk.Entry(window2, font="Arial", foreground="white", textvariable=admin_login_V)
        admin_login_entry.place(relx=0.02, rely=0.42, relwidth=0.1, relheight=0.05)

        admin_password_V = StringVar()
        admin_password_entry = ttk.Entry(window2, font="Arial", foreground="white", textvariable=admin_password_V)
        admin_password_entry.place(relx=0.02, rely=0.52, relwidth=0.1, relheight=0.05)

        admin_login_title_label = ttk.Label(window2, text='СОЗДАТЬ АДМ-АККАУНТ', font=("Arial", 10), anchor='center',
                                            background='gray10', foreground='white')
        admin_login_title_label.place(relx=0.02, rely=0.32, relwidth=0.1, relheight=0.05)
        admin_login_label = ttk.Label(window2, text='ЛОГИН', font="Arial", anchor='center', background='gray10',
                                      foreground='white')
        admin_login_label.place(relx=0.02, rely=0.37, relwidth=0.1, relheight=0.05)
        admin_password_label = ttk.Label(window2, text='ПАРОЛЬ', font="Arial", anchor='center', background='gray10',
                                         foreground='white')
        admin_password_label.place(relx=0.02, rely=0.47, relwidth=0.1, relheight=0.05)
        admin_warning_V = StringVar()
        admin_warning_label = ttk.Label(window2, textvariable=admin_warning_V, font=("Arial", 10), anchor='w',
                                        background='gray10', foreground='red')
        admin_warning_label.place(relx=0.02, rely=0.57, relwidth=0.23, relheight=0.05)

        admin_register_button = ttk.Button(window2, text='РЕГИСТРАЦИЯ', style='Admin_register.TButton', command=admin_register)
        admin_register_button.place(relx=0.02, rely=0.62, relwidth=0.1, relheight=0.05)

    root2.mainloop()


check_for_databases()

# Окно регистрации и логина
root = Tk()
root.attributes("-fullscreen", True)

style = ttk.Style()
style.theme_use('alt')
style.configure('TEntry', fieldbackground='gray15')
style.configure('Register.TButton', background='RoyalBlue1', foreground='black', font=("Arial", 15))
style.configure('Login.TButton', background='red2', foreground='black', font=("Arial", 15))
style.configure('Quit.TButton', background='red', foreground='black', font=("Arial", 15))

window = Label(background='gray10')
window.pack(fill=BOTH, expand=True, anchor='center')


login_V = StringVar()
login_entry = ttk.Entry(window, font="Arial", foreground="white", textvariable=login_V)
login_entry.place(relx=0.4, rely=0.48, relwidth=0.2, relheight=0.05)

password_V = StringVar()
password_entry = ttk.Entry(window, font="Arial", foreground="white", textvariable=password_V)
password_entry.place(relx=0.4, rely=0.68, relwidth=0.2, relheight=0.05)


title_label = ttk.Label(window, text='СИСТЕМА АККАУНТОВ', font=("Arial", 50), anchor='center', background='gray10', foreground='white')
title_label.place(relx=0, rely=0.05, relwidth=1, relheight=0.35)

login_label = ttk.Label(window, text='ЛОГИН', font="Arial", anchor='center', background='gray10', foreground='white')
login_label.place(relx=0.4, rely=0.4, relwidth=0.2, relheight=0.05)

password_label = ttk.Label(window, text='ПАРОЛЬ', font="Arial", anchor='center', background='gray10', foreground='white')
password_label.place(relx=0.4, rely=0.6, relwidth=0.2, relheight=0.05)

warning_V = StringVar()
warning_label = ttk.Label(window, textvariable=warning_V, font="Arial", anchor='center', background='gray10', foreground='red')
warning_label.place(relx=0, rely=0.3, relwidth=1, relheight=0.05)


register_button = ttk.Button(window, text='РЕГИСТР', style='Register.TButton', command=register)
register_button.place(relx=0.4, rely=0.8, relwidth=0.075, relheight=0.05)

login_button = ttk.Button(window, text='ЛОГИН', style='Login.TButton', command=login)
login_button.place(relx=0.525, rely=0.8, relwidth=0.075, relheight=0.05)

quit_frame = ttk.Label(window, background='gray8')
quit_frame.place(relx=0, rely=0, relwidth=1, relheight=0.06)

quit_button = ttk.Button(quit_frame, text='Выйти', style='Quit.TButton', command=root.destroy)
quit_button.place(relx=0.45, rely=0.1, relwidth=0.1, relheight=0.8)

root.mainloop()
